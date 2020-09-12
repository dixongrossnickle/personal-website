"""
This file is used for manipulating the values in the soccer spreadsheets
(i.e. club names)
"""

import re
import openpyxl as xl
from openpyxl.utils import get_column_letter

sheet_name = '_________'
path_to_spreadsheet = f'/Users/dixon/website/programs/19-20_test/{sheet_name}.xlsx'
teams_dict = {
    # Prem
    'Man City': 'Manchester City',
    'Liverpool': 'Liverpool',
    'Tottenham': 'Tottenham',
    'Man United': 'Manchester United',
    'Chelsea': 'Chelsea',
    'Arsenal': 'Arsenal',
    'West Ham': 'West Ham',
    'Wolves': 'Wolves',
    'Everton': 'Everton',
    'Crystal Palace': 'Crystal Palace',
    'Leicester': 'Leicester City',
    'Watford': 'Watford',
    'Bournemouth': 'Bournemouth',
    'Newcastle': 'Newcastle',
    'Brighton': 'Brighton',
    'Burnley': 'Burnley',
    'Southampton': 'Southampton',
    'Sheffield United': 'Sheffield United',
    'Norwich': 'Norwich',
    'Aston Villa': 'Aston Villa',
    # Bundesliga
    'Bayern': 'Bayern',
    'Leverkusen': 'Leverkusen',
    'Dortmund': 'Dortmund',
    "M'gladbach": "M√∂nchengladbach",
    'Frankfurt': 'Frankfurt',
    'Augsburg': 'Augsburg',
    'sseldorf': 'D√ºsseldorf',
    'Freiburg': 'Freiburg',
    'Hertha': 'Hertha',
    'Hoffenheim': 'Hoffenheim',
    'Mainz': 'Mainz',
    'Leipzig': 'RB Leipzig',
    'Koln': 'FC K√∂ln',
    'Schalke': 'Schalke',
    'Union Berlin': 'Union Berlin',
    'Wolfsburg': 'Wolfsburg',
    'Paderborn': 'Paderborn',
    'Werder Bremen': 'Werder Bremen',
    # Serie A
    'Verona': 'Verona',
    'Lazio': 'Lazio',
    'Udinese': 'Udinese',
    'Bologna': 'Bologna',
    'Brescia': 'Brescia',
    'Roma': 'Roma',
    'Atalanta': 'Atalanta',
    'Cagliari': 'Cagliari',
    'Sampdoria': 'Sampdoria',
    'Lecce': 'Lecce',
    'Genoa': 'Genoa',
    'Milan': 'AC Milan',
    'Juventus': 'Juventus',
    'Napoli': 'Napoli',
    'Sassuolo': 'Sassuolo',
    'Parma': 'Parma',
    'Torino': 'Torino',
    'Fiorentina': 'Fiorentina',
    'Spal': 'SPAL',
    'Inter': 'Inter Milan',
    # La Liga
    'Betis': 'Real Betis',
    'Granada': 'Granada',
    'Barcelona': 'Barcelona',
    'Celta': 'Celta Vigo',
    'Villarreal': 'Villarreal',
    'Eibar': 'Eibar',
    'Real Madrid': 'Real Madrid',
    'Mallorca': 'Mallorca',
    'Bilbao': 'Athletic Club',
    'Valencia': 'Valencia',
    'Getafe': 'Getafe',
    'Leganes': 'Legan√©s',
    'Alaves': 'Alav√©s',
    'Ath Madrid': 'Atl√©tico Madrid',
    'Valladolid': 'Valladolid',
    'Espanol': 'Espanyol',
    'Sevilla': 'Sevilla',
    'Levante': 'Levante',
    'Osasuna': 'Osasuna',
    'Sociedad': 'Real Sociedad',
    # Ligue 1
    'Marseille': 'Marseille',
    'Angers': 'Angers',
    'Lille': 'Lille',
    'Montpellier': 'Montpellier',
    'Nantes': 'Nantes',
    'Nice': 'Nice',
    'St Etienne': 'Saint-√âtienne',
    'Bordeaux': 'Bordeaux',
    'Lyon': 'Lyon',
    'Paris': 'PSG',
    'Reims': 'Reims',
    'Amiens': 'Amiens',
    'Brest': 'Brest',
    'Dijon': 'Dijon',
    'Metz': 'Metz',
    'Monaco': 'Monaco',
    'Rennes': 'Rennes',
    'Nimes': 'N√Æmes',
    'Strasbourg': 'Strasbourg',
    'Toulouse': 'Toulouse'
}
team_names = set()
teams_changed = set()

# Load sheet
def load_excel_sheet():
    wb = xl.load_workbook(path_to_spreadsheet)
    wb.sheetnames
    sheet = wb[sheet_name]
    num_rows = sheet.max_row
    return wb, sheet, num_rows

# Loop through every row in sheet and do something
def loop_rows(sheet, num_rows, column):
    for i in range(num_rows):
        team_name = str(sheet[f'{column}{i+1}'].value)
        for key in teams_dict:
            if re.search(key, team_name):
                sheet[f'{column}{i+1}'] = teams_dict[key]
                teams_changed.add(teams_dict[key])

# Print teams that weren't changed
def print_difference():
    difference = team_names.difference(teams_changed)
    for team_name in difference:
        print(team_name)

# Main
def main():
    for i in teams_dict.keys():
        team_names.add(teams_dict[i])

    wb, sheet, num_rows = load_excel_sheet()

    loop_rows(sheet, num_rows, 'D')
    loop_rows(sheet, num_rows, 'E')

    print_difference()

    wb.save('new_league_matches.xlsx')


main()
