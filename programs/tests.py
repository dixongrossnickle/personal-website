"""
This file is used for manipulating the values in the soccer spreadsheets
(i.e. club names)
"""

import re
import openpyxl as xl
from openpyxl.utils import get_column_letter


filename = 'NEW_PLAYERS' #(.xlsx)
sheet_name = 'FIFA 20 players'
path_to_spreadsheet = f'/Users/dixon/website/programs/football_data/{filename}.xlsx'

oldFIFA_teams_dict = {
    # Prem
    'Man City': 'Manchester City',
    'Liverpool': 'Liverpool',
    'Tottenham': 'Tottenham',
    'Man United': 'Manchester United',
    'Chelsea': 'Chelsea',
    'Arsenal': 'Arsenal',
    'West Ham': 'West Ham',
    'Wolves': 'Wolves',
    'Everton': 'Everton',
    'Crystal Palace': 'Crystal Palace',
    'Leicester': 'Leicester City',
    'Watford': 'Watford',
    'Bournemouth': 'Bournemouth',
    'Newcastle': 'Newcastle',
    'Brighton': 'Brighton',
    'Burnley': 'Burnley',
    'Southampton': 'Southampton',
    'Sheffield United': 'Sheffield United',
    'Norwich': 'Norwich',
    'Aston Villa': 'Aston Villa',
    # Bundesliga
    'Bayern': 'Bayern',
    'Leverkusen': 'Leverkusen',
    'Dortmund': 'Dortmund',
    "M'gladbach": "M√∂nchengladbach",
    'Frankfurt': 'Frankfurt',
    'Augsburg': 'Augsburg',
    'sseldorf': 'D√ºsseldorf',
    'Freiburg': 'Freiburg',
    'Hertha': 'Hertha',
    'Hoffenheim': 'Hoffenheim',
    'Mainz': 'Mainz',
    'Leipzig': 'RB Leipzig',
    'Koln': 'FC K√∂ln',
    'Schalke': 'Schalke',
    'Union Berlin': 'Union Berlin',
    'Wolfsburg': 'Wolfsburg',
    'Paderborn': 'Paderborn',
    'Werder Bremen': 'Werder Bremen',
    # Serie A
    'Verona': 'Verona',
    'Lazio': 'Lazio',
    'Udinese': 'Udinese',
    'Bologna': 'Bologna',
    'Brescia': 'Brescia',
    'Roma': 'Roma',
    'Atalanta': 'Atalanta',
    'Cagliari': 'Cagliari',
    'Sampdoria': 'Sampdoria',
    'Lecce': 'Lecce',
    'Genoa': 'Genoa',
    'Milan': 'AC Milan',
    'Juventus': 'Juventus',
    'Napoli': 'Napoli',
    'Sassuolo': 'Sassuolo',
    'Parma': 'Parma',
    'Torino': 'Torino',
    'Fiorentina': 'Fiorentina',
    'Spal': 'SPAL',
    'Inter': 'Inter Milan',
    # La Liga
    'Betis': 'Real Betis',
    'Granada': 'Granada',
    'Barcelona': 'Barcelona',
    'Celta': 'Celta Vigo',
    'Villarreal': 'Villarreal',
    'Eibar': 'Eibar',
    'Real Madrid': 'Real Madrid',
    'Mallorca': 'Mallorca',
    'Bilbao': 'Athletic Club',
    'Valencia': 'Valencia',
    'Getafe': 'Getafe',
    'Leganes': 'Legan√©s',
    'Alaves': 'Alav√©s',
    'Ath Madrid': 'Atl√©tico Madrid',
    'Valladolid': 'Valladolid',
    'Espanol': 'Espanyol',
    'Sevilla': 'Sevilla',
    'Levante': 'Levante',
    'Osasuna': 'Osasuna',
    'Sociedad': 'Real Sociedad',
    # Ligue 1
    'Marseille': 'Marseille',
    'Angers': 'Angers',
    'Lille': 'Lille',
    'Montpellier': 'Montpellier',
    'Nantes': 'Nantes',
    'Nice': 'Nice',
    'St Etienne': 'Saint-√âtienne',
    'Bordeaux': 'Bordeaux',
    'Lyon': 'Lyon',
    'Paris': 'PSG',
    'Reims': 'Reims',
    'Amiens': 'Amiens',
    'Brest': 'Brest',
    'Dijon': 'Dijon',
    'Metz': 'Metz',
    'Monaco': 'Monaco',
    'Rennes': 'Rennes',
    'Nimes': 'N√Æmes',
    'Strasbourg': 'Strasbourg',
    'Toulouse': 'Toulouse'
}

newFIFA_teams_dict = {
    'E10': "Arsenal",
    'E11': "Aston Villa",
    'E12': "Bournemouth",
    'E13': "Brighton",
    'E14': "Burnley",
    'E15': "Chelsea",
    'E16': "Crystal Palace",
    'E17': "Everton",
    'E18': "Leicester",
    'E19': "Liverpool",
    'E110': "Manchester City",
    'E111': "Manchester United",
    'E112': "Newcastle",
    'E113': "Norwich",
    'E114': "Sheffield United",
    'E115': "Southampton",
    'E116': "Tottenham",
    'E117': "Watford",
    'E118': "West Ham",
    'E119': "Wolves",
    'G10': "Augsburg",
    'G11': "Leverkusen",
    'G12': "Bayern",
    'G13': "Dortmund",
    'G14': "chengladbach",
    'G15': "Frankfurt",
    'G16': "FC Köln",
    'G17': "Düsseldorf",
    'G18': "Freiburg",
    'G19': "Hertha",
    'G110': "Hoffenheim",
    'G111': "Mainz",
    'G112': "Paderborn",
    'G113': "Leipzig",
    'G114': "Schalke",
    'G115': "Union Berlin",
    'G116': "Werder Bremen",
    'G117': "Wolfsburg",
    'I10': "AC Milan",
    'I11': "Atalanta",
    'I12': "Bologna",
    'I13': "Brescia",
    'I14': "Cagliari",
    'I15': "Fiorentina",
    'I16': "Genoa",
    'I17': "Verona",
    'I18': "Inter Milan",
    'I19': "Juventus",
    'I110': "Lazio",
    'I111': "Lecce",
    'I112': "Napoli",
    'I113': "Parma",
    'I114': "Roma",
    'I115': "SPAL",
    'I116': "Sampdoria",
    'I117': "Sassuolo",
    'I118': "Torino",
    'I119': "Udinese",
    'S10': "Alavés",
    'S11': "Athletic Club",
    'S12': "tico Madrid",
    'S13': "Barcelona",
    'S14': "Celta Vigo",
    'S15': "Eibar",
    'S16': "Espanyol",
    'S17': "Getafe",
    'S18': "Granada",
    'S19': "Leganés",
    'S110': "Levante",
    'S111': "Mallorca",
    'S112': "Osasuna",
    'S113': "Betis",
    'S114': "Real Madrid",
    'S115': "Real Sociedad",
    'S116': "Sevilla",
    'S117': "Valencia",
    'S118': "Valladolid",
    'S119': "Villarreal",
    'F10': "Amiens",
    'F11': "Angers",
    'F12': "Bordeaux",
    'F13': "Brest",
    'F14': "Dijon",
    'F15': "Lille",
    'F16': "Lyon",
    'F17': "Marseille",
    'F18': "Metz",
    'F19': "Monaco",
    'F110': "Montpellier",
    'F111': "Nantes",
    'F112': "Nice",
    'F113': "Nîmes",
    'F114': "PSG",
    'F115': "Reims",
    'F116': "Rennes",
    'F117': "Étienne",
    'F118': "Strasbourg",
    'F119': "Toulouse"
}

MATCHES_teams_dict = {
    'E10': "Arsenal",
    'E11': "Aston Villa",
    'E12': "Bournemouth",
    'E13': "Brighton",
    'E14': "Burnley",
    'E15': "Chelsea",
    'E16': "Crystal Palace",
    'E17': "Everton",
    'E18': "Leicester",
    'E19': "Liverpool",
    'E110': "Man City",
    'E111': "Man United",
    'E112': "Newcastle",
    'E113': "Norwich",
    'E114': "Sheffield United",
    'E115': "Southampton",
    'E116': "Tottenham",
    'E117': "Watford",
    'E118': "West Ham",
    'E119': "Wolves",
    'G10': "Augsburg",
    'G11': "Leverkusen",
    'G12': "Bayern",
    'G13': "Dortmund",
    'G14': "M'gladbach",
    'G15': "Frankfurt",
    'G16': "Koln",
    'G17': "Dusseldorf",
    'G18': "Freiburg",
    'G19': "Hertha",
    'G110': "Hoffenheim",
    'G111': "Mainz",
    'G112': "Paderborn",
    'G113': "RB Leipzig",
    'G114': "Schalke 04",
    'G115': "Union Berlin",
    'G116': "Werder Bremen",
    'G117': "Wolfsburg",
    'I10': "Milan",
    'I11': "Atalanta",
    'I12': "Bologna",
    'I13': "Brescia",
    'I14': "Cagliari",
    'I15': "Fiorentina",
    'I16': "Genoa",
    'I17': "Verona",
    'I18': "Inter",
    'I19': "Juventus",
    'I110': "Lazio",
    'I111': "Lecce",
    'I112': "Napoli",
    'I113': "Parma",
    'I114': "Roma",
    'I115': "Spal",
    'I116': "Sampdoria",
    'I117': "Sassuolo",
    'I118': "Torino",
    'I119': "Udinese",
    'S10': "Alaves",
    'S11': "Bilbao",
    'S12': "Ath Madrid",
    'S13': "Barcelona",
    'S14': "Celta",
    'S15': "Eibar",
    'S16': "Espanol",
    'S17': "Getafe",
    'S18': "Granada",
    'S19': "Leganes",
    'S110': "Levante",
    'S111': "Mallorca",
    'S112': "Osasuna",
    'S113': "Betis",
    'S114': "Real Madrid",
    'S115': "Sociedad",
    'S116': "Sevilla",
    'S117': "Valencia",
    'S118': "Valladolid",
    'S119': "Villarreal",
    'F10': "Amiens",
    'F11': "Angers",
    'F12': "Bordeaux",
    'F13': "Brest",
    'F14': "Dijon",
    'F15': "Lille",
    'F16': "Lyon",
    'F17': "Marseille",
    'F18': "Metz",
    'F19': "Monaco",
    'F110': "Montpellier",
    'F111': "Nantes",
    'F112': "Nice",
    'F113': "Nimes",
    'F114': "Paris",
    'F115': "Reims",
    'F116': "Rennes",
    'F117': "Etienne",
    'F118': "Strasbourg",
    'F119': "Toulouse"
}
team_names = set()
teams_changed = set()
matched_teams = {}


# Load sheet
def load_excel_sheet():
    wb = xl.load_workbook(path_to_spreadsheet)
    wb.sheetnames
    sheet = wb[sheet_name]
    num_rows = sheet.max_row
    return wb, sheet, num_rows

# Loop through every row in sheet and try to match dictionary value with team name
def loop_rows(sheet, num_rows, team_col, id_col):
    for i in range(num_rows):
        team_name = str(sheet[f'{team_col}{i+1}'].value)
        for key in newFIFA_teams_dict:
            if re.search(newFIFA_teams_dict[key], team_name):
                # sheet[f'{id_col}{i+1}'] = key
                teams_changed.add(newFIFA_teams_dict[key])
                matched_teams[newFIFA_teams_dict[key]] += 1

# Print teams that weren't changed
def print_difference():
    difference = team_names.difference(teams_changed)
    for team_name in difference:
        print(team_name)

# Main
def main():
    for key in newFIFA_teams_dict.keys():
        team_names.add(newFIFA_teams_dict[key])
        matched_teams.update( {newFIFA_teams_dict[key]: 0})
    wb, sheet, num_rows = load_excel_sheet()

    team_name_col = ''
    col_to_insert = ''
    loop_rows(sheet, num_rows, team_name_col, col_to_insert)

    print('\n-------------  Difference  -------------')
    print_difference()
    print('\n-------------  Matched  -------------')
    print('Unique # teams matched: ' + str(len(matched_teams)))
    for team in matched_teams:
        if matched_teams[team] > 35:
            print(team+":  "+str(matched_teams[team]))

    # SAVE WORKBOOK
    # wb.save('NEWNEWNEW_players.xlsx')


main()
